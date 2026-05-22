"""
scrape_visits.py
================
247Sports recruit-VISITS scraper. All data is on the visits list page itself;
NO profile clicks. Parses strictly from the verified DOM provided.

Modes:
  --season N [--teams "A,B"] [--concurrency 4] [--force]
        Scrape ONE season (the matrix passes one season per job). Writes
        per-team checkpoints checkpoints/{season}/{slug}.csv and a per-season
        consolidated season_{N}.csv (the per-season artifact).
  --combine --input-dir DIR --out-dir DIR
        Merge all season_*.csv found under DIR into ONE real .xlsx (openpyxl,
        single sheet) + ONE .csv, timestamped visits_full_YYYYMMDDTHHMMSS.*.
  --emit-seasons "all|2025|2018-2026|2018,2020"
        Print a compact JSON list of seasons (used by the workflow setup job to
        build the matrix). Needs no third-party deps.

Heights: stored CLEAN ("6-7") in memory & checkpoints' data; a leading
apostrophe is added only when SERIALIZING any CSV (so Excel won't date-convert
on open), and the final .xlsx uses a real text number-format ('@') instead, so
xlsx cells show clean "6-7" without a literal apostrophe. CSV readers strip a
single leading apostrophe on the Ht field.
"""

import argparse
import asyncio
import csv
import glob
import json
import os
import random
import re
import sys
from datetime import datetime

from team_urls import all_teams, is_fbs_in_year, team_url, team_slug

# ----------------------------------------------------------------------------
# Constants / config
# ----------------------------------------------------------------------------
# Bump SCHEMA_VERSION to invalidate ALL old checkpoints (e.g. after a parser
# change). Checkpoints whose schema != this are re-scraped.
SCHEMA_VERSION = 1

COLUMNS = [
    "247 Player ID", "Team", "High School Class", "Date of Visit",
    "Recruit Name", "High School", "City/ST", "Position", "Ht", "Wt",
    "247 HS Stars", "247 HS Rating", "247 Natl Rk", "247 HS Position Rk",
    "Status", "Commitment Date",
]

LIST_SELECTOR = 'ul.ri-page__list[data-js="recruit-list"]'
ITEM_SELECTOR = f'{LIST_SELECTOR} > li.ri-page__list-item'

USER_AGENTS = [
    ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
     "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
    ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
     "(KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"),
    ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
     "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"),
    ("Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
     "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
    ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 "
     "(KHTML, like Gecko) Version/17.4 Safari/605.1.15"),
]

# Resource/host substrings to abort for speed (no data loss).
BLOCK_SUBSTR = ["bouncex", "bounceexchange", "integralas", "il_insearch"]
BLOCK_RESOURCE_TYPES = {"image", "font", "stylesheet", "media"}

# Resilience knobs.
ABORT_THRESHOLD = 15          # consecutive team FAILUREs before aborting a job
COOLOFF_SECONDS = 30          # cool-off after a failed team
PER_TEAM_DELAY = (1.5, 4.0)   # polite randomized delay between teams (per worker)
MAX_ATTEMPTS = 3              # attempts per team (rotates UA between)
RETRY_BACKOFF = 8             # base backoff seconds (multiplied by attempt)
NAV_TIMEOUT = 45000           # goto timeout (ms)
CONTAINER_TIMEOUT = 20000     # wait_for_selector(list) timeout (ms)

# Load-More / lazy-render loop.
MAX_LOADMORE_ATTEMPTS = 50
LOADMORE_STABLE_ROUNDS = 3    # post-loop "truly gone" triple-check
LOADMORE_WAIT = (0.7, 1.4)    # seconds between scroll/click and recount
LOADMORE_FAIL_THRESHOLD = 5   # consecutive errors inside the loop before bailing

BLOCK_MARKERS = [
    "access denied", "unusual traffic", "captcha", "are you a human",
    "request blocked", "verify you are human",
]

# JS run in-page to extract every row in one pass (avoids partial DOM reads and
# is much faster than per-field round-trips). Iterates the list's DIRECT li
# children in order, carrying the date-group header down to the player rows
# beneath it. Returns raw strings; Python does the normalization.
EXTRACT_JS = r"""
() => {
  const ul = document.querySelector('ul.ri-page__list[data-js="recruit-list"]');
  if (!ul) return {error: 'no-list-container', rows: []};
  const items = ul.querySelectorAll(':scope > li.ri-page__list-item');
  const rows = [];
  let currentDate = '';
  for (const li of items) {
    if (li.classList.contains('list-header')) {
      const b = li.querySelector('b.name');
      currentDate = b ? (b.textContent || '').trim() : '';
      continue;
    }
    const row = {};
    row.visit_date_raw = currentDate;

    const nameA = li.querySelector('div.recruit a.ri-page__name-link');
    row.name = nameA ? (nameA.textContent || '').trim() : '';
    row.href = nameA ? (nameA.getAttribute('href') || '') : '';

    const meta = li.querySelector('div.recruit span.meta');
    row.meta = meta ? (meta.textContent || '').trim() : '';

    const metrics = li.querySelector('div.metrics');
    row.metrics = metrics ? (metrics.textContent || '').trim() : '';

    const ratingDiv = li.querySelector('div.rating');
    row.stars = ratingDiv
      ? ratingDiv.querySelectorAll('span.icon-starsolid.yellow').length : 0;
    const scoreEl = ratingDiv ? ratingDiv.querySelector('span.score') : null;
    row.score = scoreEl ? (scoreEl.textContent || '').trim() : '';

    const rankDiv = li.querySelector('div.rank');
    const natA = rankDiv ? rankDiv.querySelector('a.natrank') : null;
    const posA = rankDiv ? rankDiv.querySelector('a.posrank') : null;
    row.natrank = natA ? (natA.textContent || '').trim() : '';
    row.posrank = posA ? (posA.textContent || '').trim() : '';
    const natHref = natA ? (natA.getAttribute('href') || '') : '';
    const posHref = posA ? (posA.getAttribute('href') || '') : '';
    row.juco = /InstitutionGroup=JuniorCollege/i.test(natHref)
            || /InstitutionGroup=JuniorCollege/i.test(posHref);

    const posDiv = li.querySelector('div.position');
    row.position = posDiv ? (posDiv.textContent || '').trim() : '';

    const statusDiv = li.querySelector('div.status');
    let committed = '';
    let checkmark = false;
    if (statusDiv) {
      const img = statusDiv.querySelector('img[title], img[alt]');
      if (img) {
        committed = (img.getAttribute('title')
                  || img.getAttribute('alt') || '').trim();
      }
      checkmark = !!statusDiv.querySelector('b.checkmark');
    }
    row.committed_school = committed;
    row.has_checkmark = checkmark;

    rows.push(row);
  }
  return {error: '', rows: rows};
}
"""

# Try to advance any "Load More"/"Show More" control via an in-page click
# (more robust to overlays than a native click). Returns true if it clicked one.
LOADMORE_JS = r"""
() => {
  const re = /load more|show more|view more|see more/i;
  const els = document.querySelectorAll('a, button, [data-js]');
  for (const el of els) {
    const t = (el.textContent || '').trim();
    if (t && re.test(t) && el.offsetParent !== null) {
      try { el.click(); return true; } catch (e) { /* ignore */ }
    }
  }
  return false;
}
"""

# ----------------------------------------------------------------------------
# Field parsing / normalization (pure functions — easy to reason about)
# ----------------------------------------------------------------------------
_WS = re.compile(r"\s+")


def clean(s):
    if not s:
        return ""
    return _WS.sub(" ", str(s)).strip()


def is_na(v):
    return clean(v).upper() in ("", "N/A", "NA", "-", "—", "--")


def clean_na(v):
    """Return '' for N/A-like values, else cleaned text (with leading '#' stripped)."""
    if is_na(v):
        return ""
    return clean(v).lstrip("#").strip()


def extract_player_id(href):
    if not href:
        return ""
    m = re.search(r"(\d+)\s*$", href.rstrip("/"))
    return m.group(1) if m else ""


def split_meta(meta):
    """'Pleasant Valley (Chico, CA)' -> ('Pleasant Valley', 'Chico, CA')."""
    meta = clean(meta)
    if not meta:
        return "", ""
    m = re.match(r"^(?P<hs>.*?)\s*\((?P<city>[^)]*)\)\s*$", meta)
    if m:
        return clean(m.group("hs")), clean(m.group("city"))
    return meta, ""


def split_metrics(metrics):
    """'6-7 / 240' -> ('6-7', '240'). Returns CLEAN height (no apostrophe).

    The verified DOM separates height/weight with ' / ' (spaces around the
    slash). We split on THAT, not a bare '/', because a missing value renders
    as 'N/A' — splitting on a bare '/' would break 'N/A' into 'N' and 'A' and
    leak a bogus 'N' weight. clean() has already collapsed any whitespace runs
    to single spaces, so ' / ' is the reliable separator.
    """
    metrics = clean(metrics)
    if not metrics:
        return "", ""
    if " / " in metrics:
        parts = metrics.split(" / ", 1)
    else:
        # No spaced separator seen (defensive). Pull a leading height token of
        # the form D-D / D-DD; whatever remains is the weight.
        m = re.match(r"^\s*(\d{1,2}-\d{1,2})\s*/?\s*(.*)$", metrics)
        if m:
            parts = [m.group(1), m.group(2)]
        else:
            parts = [metrics, ""]
    ht = parts[0].strip() if parts else ""
    wt = parts[1].strip() if len(parts) > 1 else ""
    if is_na(ht):
        ht = ""
    if is_na(wt):
        wt = ""
    return ht, wt


def format_rank(value, juco):
    """Clean a rank; append ' (JUCO)' only on REAL values when juco=True."""
    v = clean_na(value)
    if not v:
        return ""
    return f"{v} (JUCO)" if juco else v


_DATE_COUNT = re.compile(r"\s*\(\d+\)\s*$")


def normalize_date(raw):
    """'Monday, December 9, 2024 (1)' -> '2024-12-09'. Falls back to cleaned text."""
    raw = clean(raw)
    if not raw:
        return ""
    txt = _DATE_COUNT.sub("", raw).strip()
    for fmt in ("%A, %B %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(txt, fmt).date().isoformat()
        except ValueError:
            continue
    return txt  # keep whatever the header said rather than guess


def _norm_school(s):
    return re.sub(r"[^a-z0-9]", "", clean(s).lower())


def compute_status(committed_school, has_checkmark, team):
    """'yes' iff committed (checkmark) AND the logo school == the team scraped."""
    if not has_checkmark:
        return "no"
    c = _norm_school(committed_school)
    if not c:
        return "no"
    candidates = {_norm_school(team), _norm_school(team_slug(team).replace("-", " "))}
    return "yes" if c in candidates else "no"


def normalize_row(r, team, season):
    href = r.get("href", "") or ""
    hs, city = split_meta(r.get("meta", ""))
    ht, wt = split_metrics(r.get("metrics", ""))
    juco = bool(r.get("juco", False))
    try:
        stars = min(int(r.get("stars", 0) or 0), 5)
    except (TypeError, ValueError):
        stars = 0
    return {
        "247 Player ID": extract_player_id(href),
        "Team": team,
        "High School Class": season,
        "Date of Visit": normalize_date(r.get("visit_date_raw", "")),
        "Recruit Name": clean(r.get("name", "")),
        "High School": hs,
        "City/ST": city,
        "Position": clean(r.get("position", "")),
        "Ht": ht,                                  # clean "6-7"
        "Wt": wt,
        "247 HS Stars": stars,
        "247 HS Rating": clean_na(r.get("score", "")),
        "247 Natl Rk": format_rank(r.get("natrank", ""), juco),
        "247 HS Position Rk": format_rank(r.get("posrank", ""), juco),
        "Status": compute_status(r.get("committed_school", ""),
                                 bool(r.get("has_checkmark", False)), team),
        "Commitment Date": "",   # NOT present on the visits page (see README note)
    }


# ----------------------------------------------------------------------------
# CSV (de)serialization with Excel-safe height
# ----------------------------------------------------------------------------
def _csv_row(r):
    out = dict(r)
    ht = out.get("Ht", "")
    out["Ht"] = f"'{ht}" if ht else ""     # leading apostrophe = Excel text
    return out


def _strip_ht(r):
    ht = (r.get("Ht", "") or "")
    r["Ht"] = ht.lstrip("'")
    return r


def write_csv(path, rows):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        for r in rows:
            w.writerow(_csv_row({c: r.get(c, "") for c in COLUMNS}))


# ----------------------------------------------------------------------------
# Checkpoints (per team-season) with schema + status
# ----------------------------------------------------------------------------
def checkpoint_path(season, team):
    return os.path.join("checkpoints", str(season), f"{team_slug(team)}.csv")


_HEADER_RE = re.compile(r"#\s*SCHEMA=(\d+)\s+STATUS=(\w+)")


def write_checkpoint(season, team, rows, status):
    path = checkpoint_path(season, team)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        f.write(f"# SCHEMA={SCHEMA_VERSION} STATUS={status}\n")
        if status in ("OK", "EMPTY", "NOPAGE"):
            w = csv.DictWriter(f, fieldnames=COLUMNS)
            w.writeheader()
            for r in rows:
                w.writerow(_csv_row({c: r.get(c, "") for c in COLUMNS}))


def read_checkpoint(season, team):
    """Return (rows, status) if the checkpoint is COMPLETE & current; else None
    (meaning: re-scrape). FAILED, empty, or stale-schema checkpoints -> None."""
    path = checkpoint_path(season, team)
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            first = f.readline()
            m = _HEADER_RE.search(first)
            if not m:
                return None
            schema = int(m.group(1))
            status = m.group(2).upper()
            if schema != SCHEMA_VERSION:
                return None
            if status == "FAILED":
                return None
            if status in ("EMPTY", "NOPAGE"):
                return [], status
            if status == "OK":
                reader = csv.DictReader(f)
                rows = [_strip_ht({c: row.get(c, "") for c in COLUMNS})
                        for row in reader]
                if not rows:           # OK with no rows is suspicious -> redo
                    return None
                return rows, "OK"
    except Exception:
        return None
    return None


# ----------------------------------------------------------------------------
# Page operations (lazy import of playwright happens in run_scrape)
# ----------------------------------------------------------------------------
async def _route_block(context):
    async def handler(route):
        try:
            req = route.request
            url = req.url.lower()
            if (req.resource_type in BLOCK_RESOURCE_TYPES
                    or any(b in url for b in BLOCK_SUBSTR)):
                await route.abort()
            else:
                await route.continue_()
        except Exception:
            try:
                await route.continue_()
            except Exception:
                pass
    await context.route("**/*", handler)


async def _load_more_until_done(page):
    """Scroll + click any load-more control until the row count is stable for
    LOADMORE_STABLE_ROUNDS consecutive rounds. NEVER breaks on a single error."""
    prev = -1
    stable = 0
    consec_err = 0
    for _ in range(MAX_LOADMORE_ATTEMPTS):
        if stable >= LOADMORE_STABLE_ROUNDS:
            break
        try:
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            try:
                await page.evaluate(LOADMORE_JS)
            except Exception:
                pass  # clicking is best-effort; scrolling alone may load more
            await page.wait_for_timeout(int(random.uniform(*LOADMORE_WAIT) * 1000))
            count = await page.locator(ITEM_SELECTOR).count()
            if count == prev:
                stable += 1
            else:
                stable = 0
                prev = count
            consec_err = 0
        except Exception:
            consec_err += 1
            if consec_err >= LOADMORE_FAIL_THRESHOLD:
                break
            await page.wait_for_timeout(1500)
            continue


async def scrape_team_season(browser, team, season):
    """Returns (rows, status) where status in OK / EMPTY / NOPAGE / FAILED.
    Retries up to MAX_ATTEMPTS with a fresh context + rotated UA each attempt."""
    from playwright.async_api import TimeoutError as PWTimeout

    last_err = ""
    for attempt in range(1, MAX_ATTEMPTS + 1):
        ua = random.choice(USER_AGENTS)
        context = await browser.new_context(
            user_agent=ua,
            viewport={"width": 1366, "height": 900},
            locale="en-US",
        )
        try:
            await _route_block(context)
            page = await context.new_page()
            url = team_url(team, season)
            resp = await page.goto(url, wait_until="domcontentloaded",
                                   timeout=NAV_TIMEOUT)
            status_code = resp.status if resp else 0

            if status_code == 404:
                return [], "NOPAGE"
            if status_code in (403, 429) or status_code >= 500:
                last_err = f"http {status_code}"
                raise RuntimeError(last_err)

            try:
                await page.wait_for_selector(LIST_SELECTOR,
                                             timeout=CONTAINER_TIMEOUT)
            except PWTimeout:
                # Container never appeared. Distinguish "loaded but no visits"
                # from "blocked/skeleton". Bias toward FAILED on uncertainty so
                # we never silently DROP a team (a re-fetch is cheap; data loss
                # is not).
                html = (await page.content()).lower()
                if any(mk in html for mk in BLOCK_MARKERS):
                    last_err = "blocked"
                    raise RuntimeError(last_err)
                if len(html) > 40000:        # heuristic: a real page rendered
                    return [], "EMPTY"
                last_err = "no list container / short page"
                raise RuntimeError(last_err)

            await _load_more_until_done(page)

            data = await page.evaluate(EXTRACT_JS)
            raw_rows = data.get("rows", []) if isinstance(data, dict) else []
            rows = [normalize_row(r, team, season) for r in raw_rows]
            rows = [r for r in rows if r["Recruit Name"] or r["247 Player ID"]]
            return (rows, "OK") if rows else ([], "EMPTY")

        except Exception as e:
            last_err = (str(e) or repr(e))[:200]
            if attempt < MAX_ATTEMPTS:
                await asyncio.sleep(RETRY_BACKOFF * attempt + random.uniform(0, 2))
                continue
            print(f"  [FAILED] {team} {season}: {last_err}", flush=True)
            return [], "FAILED"
        finally:
            try:
                await context.close()
            except Exception:
                pass
    return [], "FAILED"


# ----------------------------------------------------------------------------
# Orchestration (worker pool with shared consecutive-failure abort)
# ----------------------------------------------------------------------------
def resolve_teams(teams_arg, season):
    eligible = [t for t in all_teams() if is_fbs_in_year(t, season)]
    teams_arg = (teams_arg or "").strip()
    if not teams_arg or teams_arg.lower() == "all":
        return eligible
    lower_map = {t.lower(): t for t in all_teams()}
    chosen, unknown = [], []
    for name in teams_arg.split(","):
        key = name.strip().lower()
        if not key:
            continue
        if key in lower_map:
            t = lower_map[key]
            if is_fbs_in_year(t, season):
                chosen.append(t)
            else:
                print(f"  [skip] {t} not FBS in {season}", flush=True)
        else:
            unknown.append(name.strip())
    if unknown:
        print(f"  [warn] unknown team name(s) ignored: {unknown}", flush=True)
    return chosen


async def _worker(name, browser, queue, season, force, results, state,
                  lock, abort_event):
    while not abort_event.is_set():
        try:
            team = queue.get_nowait()
        except asyncio.QueueEmpty:
            return
        try:
            if not force:
                cached = read_checkpoint(season, team)
                if cached is not None:
                    rows, status = cached
                    results[team] = (rows, status, "cached")
                    print(f"  [cache] {team} {season} ({status}, {len(rows)})",
                          flush=True)
                    continue

            rows, status = await scrape_team_season(browser, team, season)
            write_checkpoint(season, team, rows, status)
            results[team] = (rows, status, "scraped")
            print(f"  [{status}] {team} {season} -> {len(rows)} rows", flush=True)

            async with lock:
                if status in ("OK", "EMPTY", "NOPAGE"):
                    state["consec"] = 0
                else:
                    state["consec"] += 1
                    if state["consec"] >= ABORT_THRESHOLD:
                        abort_event.set()
                        print(f"  [ABORT] {ABORT_THRESHOLD} consecutive failures "
                              f"-> stopping; writing partial results", flush=True)

            if status == "FAILED":
                await asyncio.sleep(COOLOFF_SECONDS)
            await asyncio.sleep(random.uniform(*PER_TEAM_DELAY))
        finally:
            queue.task_done()


def _spot_check_arkansas_2025(rows, where):
    """Print the three ground-truth Arkansas-2025 rows for eyeballing."""
    targets = ["keyshawn davila", "gavin garretson", "timothy merritt"]
    found = [r for r in rows
             if r.get("Team") == "Arkansas"
             and str(r.get("High School Class")) == "2025"
             and clean(r.get("Recruit Name", "")).lower() in targets]
    if not found:
        return
    print(f"\n--- GROUND-TRUTH SPOT CHECK ({where}, Arkansas 2025) ---", flush=True)
    for r in found:
        print(f"  {r['Recruit Name']:<20} pos={r['Position']:<5} "
              f"natl={r['247 Natl Rk']:<10} posrk={r['247 HS Position Rk']:<10} "
              f"rating={r['247 HS Rating']:<5} status={r['Status']}", flush=True)
    print("  Expect: Keyshawn Davila CB JUCO Natl 9 Pos 1; "
          "Gavin Garretson Edge rating 86 Pos 117; "
          "Timothy Merritt CB status=no (committed Tennessee).\n", flush=True)


def _print_validation(season, teams, results):
    by_status = {}
    total_rows = 0
    zero_teams, nopage_teams, failed_teams, missing = [], [], [], []
    for t in teams:
        if t not in results:
            missing.append(t)
            continue
        rows, status, _src = results[t]
        by_status[status] = by_status.get(status, 0) + 1
        total_rows += len(rows)
        if status == "OK" and len(rows) == 0:
            zero_teams.append(t)
        if status == "EMPTY":
            zero_teams.append(t)
        if status == "NOPAGE":
            nopage_teams.append(t)
        if status == "FAILED":
            failed_teams.append(t)

    print("\n" + "=" * 60, flush=True)
    print(f"VALIDATION — season {season}", flush=True)
    print(f"  teams attempted : {len(teams)}", flush=True)
    print(f"  by status       : {by_status}", flush=True)
    print(f"  total rows      : {total_rows}", flush=True)
    if missing:
        print(f"  MISSING (queue not finished / aborted): {missing}", flush=True)
    if failed_teams:
        print(f"  FAILED (will re-scrape next run): {failed_teams}", flush=True)
    if nopage_teams:
        print(f"  NO PAGE / 404 (check slug & FBS year): {nopage_teams}", flush=True)
    if zero_teams:
        print(f"  EMPTY / 0 visits (verify a couple are truly empty): "
              f"{zero_teams}", flush=True)
    print("=" * 60 + "\n", flush=True)


async def run_scrape(season, teams_arg, concurrency, force):
    from playwright.async_api import async_playwright

    teams = resolve_teams(teams_arg, season)
    print(f"Scraping season {season}: {len(teams)} teams, "
          f"concurrency={concurrency}, force={force}", flush=True)

    queue = asyncio.Queue()
    for t in teams:
        queue.put_nowait(t)

    results = {}
    state = {"consec": 0}
    lock = asyncio.Lock()
    abort_event = asyncio.Event()

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage",
                  "--disable-blink-features=AutomationControlled"],
        )
        try:
            workers = [
                asyncio.create_task(
                    _worker(f"w{i}", browser, queue, season, force,
                            results, state, lock, abort_event))
                for i in range(max(1, concurrency))
            ]
            await asyncio.gather(*workers)
        finally:
            await browser.close()

    # Consolidate this season's rows into the per-season artifact.
    all_rows = []
    for t in teams:
        if t in results:
            all_rows.extend(results[t][0])
    out_csv = f"season_{season}.csv"
    write_csv(out_csv, all_rows)
    print(f"Wrote {out_csv}: {len(all_rows)} rows", flush=True)

    _print_validation(season, teams, results)
    if season == 2025:
        _spot_check_arkansas_2025(all_rows, "live run")

    if abort_event.is_set():
        # Non-zero exit so the run is visibly flagged, but artifacts are written.
        print("Exiting non-zero: aborted on sustained failures.", flush=True)
        sys.exit(1)


# ----------------------------------------------------------------------------
# Combine -> one real xlsx + one csv
# ----------------------------------------------------------------------------
def parse_seasons(spec):
    spec = (spec or "").strip().lower()
    from team_urls import DEFAULT_SEASONS
    if not spec or spec == "all":
        return list(DEFAULT_SEASONS)
    years = set()
    for part in spec.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            a, b = int(a), int(b)
            if a > b:
                a, b = b, a
            years.update(range(a, b + 1))
        else:
            years.add(int(part))
    years = sorted(y for y in years if 2000 <= y <= 2035)
    return years or list(DEFAULT_SEASONS)


def _read_all_season_csvs(input_dir):
    paths = sorted(glob.glob(os.path.join(input_dir, "**", "season_*.csv"),
                             recursive=True))
    if not paths:
        paths = sorted(glob.glob(os.path.join(input_dir, "season_*.csv")))
    rows, seen = [], set()
    for p in paths:
        with open(p, "r", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                r = _strip_ht({c: row.get(c, "") for c in COLUMNS})
                key = tuple(r.get(c, "") for c in COLUMNS)
                if key in seen:
                    continue
                seen.add(key)
                rows.append(r)
    return rows, paths


def _write_xlsx(path, rows):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "visits"
    ws.append(COLUMNS)
    for r in rows:
        ws.append([r.get(c, "") for c in COLUMNS])

    # Header style + freeze.
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    # Force TEXT format on Player ID (col 1) and Ht (col 9) so Excel never
    # reformats them (no scientific notation; no "6-7" -> date).
    id_letter = get_column_letter(COLUMNS.index("247 Player ID") + 1)
    ht_letter = get_column_letter(COLUMNS.index("Ht") + 1)
    for col_letter in (id_letter, ht_letter):
        for cell in ws[col_letter][1:]:
            cell.number_format = "@"

    widths = {
        "247 Player ID": 12, "Team": 18, "High School Class": 10,
        "Date of Visit": 13, "Recruit Name": 22, "High School": 26,
        "City/ST": 20, "Position": 9, "Ht": 7, "Wt": 6,
        "247 HS Stars": 6, "247 HS Rating": 9, "247 Natl Rk": 12,
        "247 HS Position Rk": 14, "Status": 8, "Commitment Date": 14,
    }
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 12)

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)


def do_combine(input_dir, out_dir):
    rows, paths = _read_all_season_csvs(input_dir)
    print(f"Combine: read {len(paths)} season file(s), {len(rows)} unique rows",
          flush=True)
    rows.sort(key=lambda r: (str(r.get("Team", "")),
                             str(r.get("High School Class", "")),
                             str(r.get("Date of Visit", "")),
                             str(r.get("Recruit Name", ""))))
    ts = datetime.now().strftime("%Y%m%dT%H%M%S")
    os.makedirs(out_dir, exist_ok=True)
    xlsx_path = os.path.join(out_dir, f"visits_full_{ts}.xlsx")
    csv_path = os.path.join(out_dir, f"visits_full_{ts}.csv")
    _write_xlsx(xlsx_path, rows)
    write_csv(csv_path, rows)
    print(f"Wrote {xlsx_path}", flush=True)
    print(f"Wrote {csv_path}", flush=True)

    # Quick combined summary + ground-truth spot check.
    per_team = {}
    for r in rows:
        per_team[r.get("Team", "")] = per_team.get(r.get("Team", ""), 0) + 1
    print(f"  teams represented : {len(per_team)}", flush=True)
    print(f"  total rows        : {len(rows)}", flush=True)
    _spot_check_arkansas_2025(rows, "combined output")


# ----------------------------------------------------------------------------
# Entry point
# ----------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="247Sports visits-list scraper")
    ap.add_argument("--season", type=int, help="single season to scrape")
    ap.add_argument("--teams", default="", help='comma list or "" for all')
    ap.add_argument("--concurrency", type=int, default=4)
    ap.add_argument("--force", action="store_true", help="ignore checkpoints")
    ap.add_argument("--combine", action="store_true")
    ap.add_argument("--input-dir", default="all_seasons")
    ap.add_argument("--out-dir", default="output")
    ap.add_argument("--emit-seasons", default=None,
                    help='print JSON season list (e.g. "2018-2026")')
    args = ap.parse_args()

    if args.emit_seasons is not None:
        print(json.dumps(parse_seasons(args.emit_seasons)))
        return
    if args.combine:
        do_combine(args.input_dir, args.out_dir)
        return
    if args.season is None:
        print("ERROR: --season is required (or use --combine / --emit-seasons)",
              file=sys.stderr)
        sys.exit(2)
    asyncio.run(run_scrape(args.season, args.teams, args.concurrency, args.force))


if __name__ == "__main__":
    main()
